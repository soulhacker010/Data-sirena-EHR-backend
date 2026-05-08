"""
Load the Office Ally payer directory from payers.xlsx into the Payer table.

Usage:
    python manage.py seed_payers                    # Load from repo-root payers.xlsx
    python manage.py seed_payers --file path/to.xlsx
    python manage.py seed_payers --clear            # Wipe and reload

The xlsx has one row per (payer, transaction-type) combo, so we aggregate
by payer_id: one Payer row per payer with boolean flags for which transaction
types it supports (837P, eligibility, ERA).
"""
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.billing.models import Payer


TXN_837P = 'Professional Claims 837P'
TXN_ELIGIBILITY = 'Eligibility 270 / 271'
TXN_ERA = 'Remits 835'


class Command(BaseCommand):
    help = 'Seed the Payer table from the Office Ally payers.xlsx directory.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default=None,
            help='Path to the payers.xlsx (defaults to <repo-root>/payers.xlsx).',
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Delete all existing Payer rows before loading.',
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError(
                'openpyxl is required. Run: pip install openpyxl'
            ) from exc

        xlsx_path = Path(options['file']) if options['file'] else self._default_path()
        if not xlsx_path.exists():
            raise CommandError(f'File not found: {xlsx_path}')

        self.stdout.write(f'Loading payers from: {xlsx_path}')

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        # Aggregate rows by payer_id — one xlsx row per (payer, transaction-type)
        aggregated: dict[str, dict] = {}
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, payer_id, txn, available, _non_par, enrollment, *_rest = row
            notes = row[9] if len(row) > 9 else ''

            name = (name or '').strip()
            payer_id = (payer_id or '').strip()
            txn = (txn or '').strip()

            if not name or not payer_id:
                skipped += 1
                continue

            entry = aggregated.setdefault(payer_id, {
                'name': name,
                'payer_id': payer_id,
                'transactions': set(),
                'available': False,
                'enrollment_required': False,
                'notes': '',
            })

            if txn:
                entry['transactions'].add(txn)
            if (available or '').strip().lower() == 'yes':
                entry['available'] = True
            if (enrollment or '').strip().lower() == 'yes':
                entry['enrollment_required'] = True
            if notes and not entry['notes']:
                entry['notes'] = str(notes).strip()

        self.stdout.write(f'Aggregated {len(aggregated)} unique payers '
                          f'({skipped} rows skipped for missing name/ID).')

        with transaction.atomic():
            if options['clear']:
                deleted, _ = Payer.objects.all().delete()
                self.stdout.write(self.style.WARNING(f'Cleared {deleted} existing Payer rows.'))

            created = 0
            updated = 0
            for payer_id, data in aggregated.items():
                txns = data['transactions']
                defaults = {
                    'name': data['name'],
                    'transaction_type': ', '.join(sorted(txns)),
                    'available': data['available'],
                    'enrollment_required': data['enrollment_required'],
                    'supports_837p': TXN_837P in txns,
                    'supports_eligibility': TXN_ELIGIBILITY in txns,
                    'supports_era': TXN_ERA in txns,
                    'notes': data['notes'],
                }
                _, was_created = Payer.objects.update_or_create(
                    payer_id=payer_id,
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(self.style.SUCCESS(
            f'[OK] Payers loaded — created: {created}, updated: {updated}'
        ))

    def _default_path(self) -> Path:
        # BASE_DIR is sirena-backend/; payers.xlsx lives in the repo root (parent)
        return Path(settings.BASE_DIR).parent / 'payers.xlsx'
